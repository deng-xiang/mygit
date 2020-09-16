import requests
import openpyxl
import time
import re
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment



def check(rul):
    headers = {

        'Accept': '*/*',
        'Accept-Encoding': 'gzip, deflate',
        'Connection': 'close',
        'Host': 'httpbin.org',
        'referer': 'https://xiaoyuan.lagou.com/',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/49.0.2623.221 Safari/537.36 SE 2.X MetaSr 1.0',

    }
    try:
        tmpstr=rul.replace('http://','')

        #print(re.sub(r'/$','',tmpstr))
        headers['Host']=re.sub(r'/$','',tmpstr)
        print(headers['Host'])
        code=requests.get(rul,headers=headers,timeout=2).status_code
    except Exception as err:
        err='异常'
        return  err
    return  code

###
def initweb(qurl,aurl,wbook,sheetname):
    ws=wbook.create_sheet(sheetname)
    ws['A1'].value='网站域名'
    ws['A2'].value='网站标题'
    ws['A3'].value='地理位置'
    ws['A4'].value='IP地址'
    ws['A5'].value='操作系统'
    ws['A6'].value='服务器技术'
    ws['A7'].value='数据库'
    ws['A8'].value='域名到期'
    ws['A9'].value='ALexa排名'
    ws['A10'].value='ICP备案'

    url = aurl
    postdata = {
        "q": qurl
    }
    r = requests.post(url, postdata)
    #print (r.text)

    soup = BeautifulSoup(r.text, 'html.parser')
    print(qurl)
    if 'http://www.glswxgj.gov.cn/' in qurl:
    #网址是否能访问
        webadress= soup.find('div', attrs={'id':'tipinfo'})
        print(webadress)
        webadress=webadress.find('div', attrs={'class': 'col-red lh30 fz14'})
        print(webadress)
        print('==========')
        if '获取不到' in webadress.string:
            print('网址不可访问')
            return False


    # 网站标题

    url = soup.find('div', attrs={'class': 'ball color-63'})
    #print("网站标题:" + url.string)
    ws['B2'].value=str(url.string)
    #网站域名
    ws['B1'].value=qurl
    #操作系统
    ws['B5'].value='未检测到'
    #数据库
    ws['B7'].value = '未检测到'
    #域名到期
    exptire_date=soup.find('table', attrs={'class':'_chinaz-seo-newt'})
    exptire_date=exptire_date.find_all('tr')
    exptire_date=exptire_date[2]
    exptire_date=exptire_date.find_all('span')
    exptire_date=exptire_date[2]
    exptire_date=exptire_date.find('a')

    ws['B8'].value=exptire_date.string.split('为')[1].replace(')','')
    # 备案信息
    document = soup.find('td', attrs={'class': '_chinaz-seo-newtc _chinaz-seo-newh40'})
    # print(document.span.a)
    document = document.find('span')
    document = document.find('i')
    document = document.find('a')
    #print("备案号:" + document.string)
    ws['B10'].value=document.string
    # IP
    IP = soup.find('td', attrs={'class': '_chinaz-seo-newh78 _chinaz-seo-newinfo'})

    IP = IP.find('div', attrs={'class': 'pb5'})
    IP = IP.find('span')
    IP = IP.find('i')

    IP = IP.find('a')
    if IP == None:
        print("网站无法获取IP")
        return False


    ip1 = IP.string.split('[')[0]
    addr = IP.string.split('[')[1].replace(']', '')
    #print("IP:" + ip1)
    ws['B4'].value=ip1
    #print("物理位置:" + addr)
    ws['B3'].value=addr





    # 服务器类型
    try:

        machinetype = soup.find('div',attrs={'class':'Manin01List03 clearfix _chinaz-seo-new11'})
        #print(machinetype)
        machinetype = machinetype.find('ul', attrs={'class': 'MaLi03List fl'})
        machinetype = machinetype.find_all('div', attrs={'class': 'MaLi03Row w180'})

        #print('服务器类型:' + machinetype[2].string)
        if '-' in machinetype[2].string:
            ws['B6'].value='未检测到'
        else:
            ws['B6'].value=machinetype[2].string.strip()
    except Exception as err:
        ws['B6'].value='未检测到'
    # ALEXA世界排名
    alexno = soup.find('i', attrs={'class': 'alexarank color-63'})
    alexno = alexno.find('a')
    if  '-' in alexno.string :
        print('Alexa世界排名:' + '未检测到')
        ws['B9'].value='未检测到'
    else:
        print('Alexa世界排名:' + alexno.string)
        ws['B9'].value = alexno.string

    return  True
###

file_name='weburl.xlsx'
wb=openpyxl.load_workbook(file_name)
sheet=wb['web']

wirte_file='init.xlsx'
wt=openpyxl.Workbook()



for row in range(2,sheet.max_row+1):
    checkurl=sheet['C'+str(row)].value
    #sheet['D'+str(row)].value=check( checkurl )
    #time.sleep(10)
    print(str(sheet['B'+str(row)].value) +':'+ str(sheet['D'+str(row)].value) )
    #if str(sheet['D'+str(row)].value)=='200':
    initweb(checkurl,'http://seo.chinaz.com',wt,sheet['B'+str(row)].value)




################绘制表格
#定义每个栏的规格 水平居中 垂直居中 自动换行
align = Alignment(horizontal='center',vertical='center',wrap_text=True)
medium = Side(border_style="medium", color="000000")
border = Border(top=medium, left=medium, right=medium, bottom=medium)

sheetnames=wt.sheetnames
print (sheetnames)
for sheetname in sheetnames:
    sheet=wt[sheetname]
    for row in sheet.rows:
        for cell in row:
            #sheet_nt[row].alignment = align
            cell.border=border
            cell.alignment=align
            #cell.aligment=align
        sheet.column_dimensions['B'].width=40.0
        sheet.column_dimensions['A'].width=15.0

wb.save(file_name)
sheet=wt['Sheet']
wt.remove(sheet)
wt.save(wirte_file)















